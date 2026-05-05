"""
US Aylık Model Backtest — 10 Strateji
======================================
OOT dönemi boyunca her ay 10 strateji için portföy simülasyonu:
  Mom_Pure/NoDualHot/LowVol/SectorDiv — Momentum × 4 seçim tipi
  Bir_Pure/NoDualHot/LowVol/SectorDiv — Birleşik (M1+M2) × 4 seçim tipi
  Reversal — Saf reversal (M2) top-5
  Karma    — 4 Momentum (pure top-4) + 1 Reversal

Kullanım:
    python backtest.py
    python backtest.py --start 2024-01-01 --capital 100

Çıktı:
    output/backtest_YYYYMMDD.xlsx  (4 sayfa: Aylık_Özet / Hisse_Detay / Performans_Özeti / Aylık_Adaylar)
    output/backtest_YYYYMMDD.png
"""

import argparse
import pickle
import sys
import warnings
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.gridspec import GridSpec

import numpy as np
import pandas as pd
import yfinance as yf

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).parent))
from pipeline import apply_winsorizer

BASE_DIR   = Path(__file__).parent
SAVE_DIR   = BASE_DIR / "model" / "saved"
OUTPUT_DIR = BASE_DIR / "output"

OOT_START_DEFAULT = "2024-01-01"
TOP_N             = 5
INIT_CASH_DEFAULT = 100.0

W_MOMENTUM        = 0.5
W_REVERSAL        = 0.5
HOT_MOM_THRESHOLD = 0.95
CAND_THRESHOLD    = 0.60
LOWVOL_POOL_MULT  = 5

STRATEGY_DEFS = {
    "Mom_Pure":      {"score": "m1",    "sel": "pure"},
    "Mom_NoDualHot": {"score": "m1",    "sel": "nodualHot"},
    "Mom_LowVol":    {"score": "m1",    "sel": "lowvol"},
    "Mom_SectorDiv": {"score": "m1",    "sel": "sectorDiv"},
    "Bir_Pure":      {"score": "bir",   "sel": "pure"},
    "Bir_NoDualHot": {"score": "bir",   "sel": "nodualHot"},
    "Bir_LowVol":    {"score": "bir",   "sel": "lowvol"},
    "Bir_SectorDiv": {"score": "bir",   "sel": "sectorDiv"},
    "Reversal":      {"score": "m2",    "sel": "pure"},
    "Karma":         {"score": "karma", "sel": "karma"},
}

SCORE_COL  = {"m1": "rank_m1", "m2": "rank_m2", "bir": "rank_bir", "karma": "rank_m1"}
KEY_STRATS = ["Bir_Pure", "Mom_Pure", "Reversal", "Karma"]


# ── Yükleme ──────────────────────────────────────────────────────────────
def load_art(name: str) -> dict | None:
    path = SAVE_DIR / name
    if not path.exists():
        print(f"  UYARI: {name} bulunamadı.")
        return None
    with open(path, "rb") as f:
        return pickle.load(f)


def load_latest_parquet() -> pd.DataFrame:
    files = sorted(OUTPUT_DIR.glob("features_monthly*.parquet"))
    if not files:
        raise FileNotFoundError(f"Parquet bulunamadı: {OUTPUT_DIR}")
    df = pd.read_parquet(files[-1])
    df["month_end"] = pd.to_datetime(df["month_end"])
    print(f"Parquet yüklendi: {files[-1].name}  ({len(df):,} satır)")
    return df


def add_ret_ranks(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["ret_1m", "ret_2m"]:
        rank_col = col + "_rank"
        if col in df.columns and rank_col not in df.columns:
            df[rank_col] = df.groupby("month_end")[col].rank(pct=True)
    return df


# ── Tahmin ───────────────────────────────────────────────────────────────
def predict_scores(art: dict, month_df: pd.DataFrame) -> pd.Series:
    df = month_df.copy()
    df["sector_enc"] = art["label_enc"].transform(
        df["sector"].fillna("Unknown").astype(str)
    )
    df = apply_winsorizer(df, art["bounds"])
    for c in art["feat_cols"]:
        if c not in df.columns:
            df[c] = 0.0
    X = df[art["feat_cols"]].values
    w = art["weights"]
    preds = (w[0] * art["xgb_model"].predict(X) +
             w[1] * art["lgb_model"].predict(X) +
             w[2] * art["cat_model"].predict(X))
    return pd.Series(preds, index=month_df.index)


# ── Seçim ────────────────────────────────────────────────────────────────
def _select_nodual_hot(md: pd.DataFrame, score_col: str, n: int = 5) -> pd.DataFrame:
    hot_1m = md["ret_1m_rank"] > HOT_MOM_THRESHOLD if "ret_1m_rank" in md.columns else pd.Series(False, index=md.index)
    hot_2m = md["ret_2m_rank"] > HOT_MOM_THRESHOLD if "ret_2m_rank" in md.columns else pd.Series(False, index=md.index)
    cands  = md[~(hot_1m & hot_2m)]
    return (cands if len(cands) >= n else md).nlargest(n, score_col)


def _select_low_vol(md: pd.DataFrame, score_col: str, n: int = 5) -> pd.DataFrame:
    pool = md.nlargest(n * LOWVOL_POOL_MULT, score_col)
    if "hvol_63d" in pool.columns and pool["hvol_63d"].notna().sum() >= n:
        return pool.nsmallest(n, "hvol_63d")
    return pool.head(n)


def _select_sector_div(md: pd.DataFrame, score_col: str, n: int = 5,
                       max_per_sector: int = 2) -> pd.DataFrame:
    sector_counts: dict = {}
    selected = []
    for _, row in md.sort_values(score_col, ascending=False).iterrows():
        sec = str(row.get("sector") or "Unknown")
        if sector_counts.get(sec, 0) < max_per_sector:
            selected.append(row)
            sector_counts[sec] = sector_counts.get(sec, 0) + 1
        if len(selected) == n:
            break
    if len(selected) < n:
        taken = {r["ticker"] for r in selected}
        fill  = md[~md["ticker"].isin(taken)].nlargest(n - len(selected), score_col)
        return pd.concat([pd.DataFrame(selected), fill]).sort_values(score_col, ascending=False)
    return pd.DataFrame(selected).sort_values(score_col, ascending=False)


def select_picks(md: pd.DataFrame, score_col: str, sel: str, n: int = 5) -> pd.DataFrame:
    if sel == "nodualHot":
        return _select_nodual_hot(md, score_col, n)
    if sel == "lowvol":
        return _select_low_vol(md, score_col, n)
    if sel == "sectorDiv":
        return _select_sector_div(md, score_col, n)
    return md.nlargest(n, score_col)


def select_karma(md: pd.DataFrame, n: int = 5) -> pd.DataFrame:
    """4 Momentum (pure top-4) + 1 Reversal."""
    top4 = md.nlargest(4, "rank_m1")
    excl = set(top4["ticker"])
    cand = md[~md["ticker"].isin(excl)]
    if "rank_m2" not in md.columns or cand.empty:
        return top4.head(n)
    top1_rev = cand.nlargest(1, "rank_m2")
    return pd.concat([top4, top1_rev]).reset_index(drop=True).head(n)


# ── S&P500 ───────────────────────────────────────────────────────────────
def fetch_sp500_returns(month_ends) -> dict:
    months = sorted(pd.Timestamp(m) for m in month_ends)
    if len(months) < 2:
        return {}
    start_s = (months[0]  - pd.Timedelta(days=10)).strftime("%Y-%m-%d")
    end_s   = (months[-1] + pd.Timedelta(days=45)).strftime("%Y-%m-%d")
    try:
        raw = yf.download("^GSPC", start=start_s, end=end_s,
                          auto_adjust=True, progress=False)
        if raw.empty:
            return {}
        prices = raw["Close"].squeeze()
        prices.index = pd.to_datetime(prices.index)
    except Exception as e:
        print(f"  S&P500 alınamadı: {e}")
        return {}

    sp_rets = {}
    for i, m in enumerate(months[:-1]):
        nxt = months[i + 1]
        before_m = prices.index[prices.index <= m]
        before_n = prices.index[prices.index <= nxt]
        if len(before_m) == 0 or len(before_n) == 0:
            continue
        p0 = float(prices.loc[before_m[-1]])
        p1 = float(prices.loc[before_n[-1]])
        sp_rets[m] = (p1 - p0) / p0
    return sp_rets


# ── Backtest ─────────────────────────────────────────────────────────────
def run_backtest(df: pd.DataFrame,
                 art1: dict,
                 art2: dict | None,
                 oot_start: str,
                 init_cash: float):
    oot    = df[df["month_end"] >= pd.Timestamp(oot_start)].copy()
    months = sorted(oot["month_end"].unique())

    print("S&P500 verisi çekiliyor...")
    sp_rets_map = fetch_sp500_returns(months)

    portfolios = {s: init_cash for s in STRATEGY_DEFS}
    sp_port    = init_cash

    monthly_rows   = []
    detail_rows    = []
    candidate_rows = []

    for month in months:
        md       = oot[oot["month_end"] == month].copy()
        md_valid = md.dropna(subset=["target_ret_1m"])
        if len(md_valid) < TOP_N:
            continue

        # Skorlar
        if art1:
            try:
                s1 = predict_scores(art1, md_valid)
                md_valid["rank_m1"] = s1.rank(pct=True).values
            except Exception as e:
                print(f"  M1 hata ({month.strftime('%Y-%m')}): {e}")
                continue

        if art2:
            try:
                s2 = predict_scores(art2, md_valid)
                md_valid["rank_m2"] = s2.rank(pct=True).values
            except Exception as e:
                print(f"  M2 hata ({month.strftime('%Y-%m')}): {e}")

        if "rank_m1" in md_valid.columns and "rank_m2" in md_valid.columns:
            md_valid["rank_bir"] = W_MOMENTUM * md_valid["rank_m1"] + W_REVERSAL * md_valid["rank_m2"]
        elif "rank_m1" in md_valid.columns:
            md_valid["rank_bir"] = md_valid["rank_m1"]

        sp_ret = sp_rets_map.get(month, np.nan)
        new_sp = sp_port * (1 + sp_ret) if pd.notna(sp_ret) else sp_port

        row = {
            "ay":           month,
            "sp500_getiri": sp_ret,
            "sp500_deger":  new_sp,
        }

        for strat, sdef in STRATEGY_DEFS.items():
            sc = SCORE_COL.get(sdef["score"], "rank_m1")
            if sc not in md_valid.columns:
                row[f"{strat}_getiri"] = np.nan
                row[f"{strat}_deger"]  = portfolios[strat]
                continue

            if sdef["sel"] == "karma" and "rank_m2" in md_valid.columns:
                picks = select_karma(md_valid, n=TOP_N)
            else:
                picks = select_picks(md_valid, sc, sdef["sel"], TOP_N)

            rets     = picks["target_ret_1m"].fillna(0).values
            port_ret = float(rets.mean())
            old_val  = portfolios[strat]
            new_val  = old_val * (1 + port_ret)
            portfolios[strat] = new_val

            row[f"{strat}_getiri"] = port_ret
            row[f"{strat}_deger"]  = new_val

            per_stock = old_val / TOP_N
            for _, r in picks.iterrows():
                ret_val = r.get("target_ret_1m", np.nan)
                hot_1m  = "✓" if pd.notna(r.get("ret_1m_rank")) and r["ret_1m_rank"] > HOT_MOM_THRESHOLD else ""
                detail_rows.append({
                    "ay":       month,
                    "strateji": strat,
                    "hisse":    r["ticker"],
                    "sektor":   r.get("sector", ""),
                    "rank_m1":  round(float(r["rank_m1"]),  4) if pd.notna(r.get("rank_m1"))  else np.nan,
                    "rank_m2":  round(float(r["rank_m2"]),  4) if pd.notna(r.get("rank_m2"))  else np.nan,
                    "rank_bir": round(float(r["rank_bir"]), 4) if pd.notna(r.get("rank_bir")) else np.nan,
                    "hot_1m":   hot_1m,
                    "getiri":   ret_val,
                    "giris":    round(per_stock, 4),
                    "cikis":    round(per_stock * (1 + (ret_val if pd.notna(ret_val) else 0)), 4),
                })

        # Aday havuzu (birleşik rank > %60)
        if "rank_bir" in md_valid.columns:
            cands = md_valid[md_valid["rank_bir"] > CAND_THRESHOLD].sort_values(
                "rank_bir", ascending=False
            )
            for sira, (_, r) in enumerate(cands.iterrows(), 1):
                candidate_rows.append({
                    "ay":       month,
                    "sira":     sira,
                    "hisse":    r["ticker"],
                    "rank_m1":  round(float(r["rank_m1"]),  4) if pd.notna(r.get("rank_m1"))  else np.nan,
                    "rank_m2":  round(float(r["rank_m2"]),  4) if pd.notna(r.get("rank_m2"))  else np.nan,
                    "rank_bir": round(float(r["rank_bir"]), 4),
                    "getiri":   r.get("target_ret_1m", np.nan),
                    "hot_1m":   "✓" if pd.notna(r.get("ret_1m_rank")) and r["ret_1m_rank"] > HOT_MOM_THRESHOLD else "",
                })

        sp_port = new_sp
        monthly_rows.append(row)

    return (pd.DataFrame(monthly_rows),
            pd.DataFrame(detail_rows),
            pd.DataFrame(candidate_rows))


# ── Metrikler ─────────────────────────────────────────────────────────────
def compute_metrics(monthly: pd.DataFrame, init_cash: float) -> dict:
    results = {}
    if monthly.empty:
        return results

    n        = len(monthly)
    sp_rets  = monthly["sp500_getiri"].fillna(0) if "sp500_getiri" in monthly.columns else pd.Series([0.0]*n)
    sp_vals  = monthly["sp500_deger"]            if "sp500_deger"  in monthly.columns else pd.Series([init_cash]*n)
    sp_total = sp_vals.iloc[-1] / init_cash - 1
    sp_ann   = (1 + sp_total) ** (12 / n) - 1 if n > 0 else np.nan

    results["__sp500__"] = {
        "Toplam Getiri":  f"%{sp_total*100:.1f}",
        "Yıllık Getiri":  f"%{sp_ann*100:.1f}" if pd.notna(sp_ann) else "N/A",
        "Sharpe":         "—",
        "Max Drawdown":   "—",
        "Kazanma Oranı":  "—",
        "Alpha (yıllık)": "—",
        "Info Ratio":     "—",
        "Son Değer ($)":  f"${sp_vals.iloc[-1]:.2f}",
        "Başlangıç":      monthly["ay"].iloc[0].strftime("%Y-%m"),
        "Bitiş":          monthly["ay"].iloc[-1].strftime("%Y-%m"),
        "Ay Sayısı":      str(n),
    }

    for strat in STRATEGY_DEFS:
        ret_col = f"{strat}_getiri"
        val_col = f"{strat}_deger"
        if ret_col not in monthly.columns:
            continue
        rets    = monthly[ret_col].fillna(0)
        vals    = monthly[val_col]
        total_r = vals.iloc[-1] / init_cash - 1
        ann_r   = (1 + total_r) ** (12 / n) - 1 if n > 0 else np.nan
        sharpe  = rets.mean() / rets.std() * np.sqrt(12) if rets.std() > 0 else np.nan
        roll    = vals.cummax()
        max_dd  = ((vals - roll) / roll).min()
        win_r   = (rets > 0).mean()
        exc     = rets.values - sp_rets.values
        ir      = exc.mean() / exc.std() * np.sqrt(12) if exc.std() > 0 else np.nan
        alpha   = ann_r - sp_ann if pd.notna(sp_ann) and pd.notna(ann_r) else np.nan

        results[strat] = {
            "Toplam Getiri":  f"%{total_r*100:.1f}",
            "Yıllık Getiri":  f"%{ann_r*100:.1f}" if pd.notna(ann_r) else "N/A",
            "Sharpe":         f"{sharpe:.2f}" if pd.notna(sharpe) else "N/A",
            "Max Drawdown":   f"%{max_dd*100:.1f}",
            "Kazanma Oranı":  f"%{win_r*100:.1f}",
            "Alpha (yıllık)": f"%{alpha*100:.1f}" if pd.notna(alpha) else "N/A",
            "Info Ratio":     f"{ir:.2f}" if pd.notna(ir) else "N/A",
            "Son Değer ($)":  f"${vals.iloc[-1]:.2f}",
            "Başlangıç":      monthly["ay"].iloc[0].strftime("%Y-%m"),
            "Bitiş":          monthly["ay"].iloc[-1].strftime("%Y-%m"),
            "Ay Sayısı":      str(n),
        }

    return results


def _build_summary_df(metrics: dict) -> pd.DataFrame:
    metric_keys = [
        "Toplam Getiri", "Yıllık Getiri", "Sharpe", "Max Drawdown",
        "Kazanma Oranı", "Alpha (yıllık)", "Info Ratio", "Son Değer ($)",
        "Başlangıç", "Bitiş", "Ay Sayısı",
    ]
    all_strats = [s for s in STRATEGY_DEFS if s in metrics]
    rows = []
    for mk in metric_keys:
        row = {"Metrik": mk}
        for s in all_strats:
            row[s] = metrics.get(s, {}).get(mk, "N/A")
        row["S&P500"] = metrics.get("__sp500__", {}).get(mk, "N/A")
        rows.append(row)
    return pd.DataFrame(rows)


# ── Excel çıktısı ─────────────────────────────────────────────────────────
def save_excel(monthly: pd.DataFrame,
               detail: pd.DataFrame,
               candidates: pd.DataFrame,
               metrics: dict,
               art1: dict,
               art2: dict | None,
               init_cash: float,
               path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def pct(v): return f"%{v*100:.2f}" if pd.notna(v) else ""
    def usd(v): return f"${v:.2f}"     if pd.notna(v) else ""
    def fmt4(v): return f"{v:.4f}"     if pd.notna(v) else ""

    H_FILL  = PatternFill("solid", fgColor="1F4E79")
    G_FILL  = PatternFill("solid", fgColor="C6EFCE")
    R_FILL  = PatternFill("solid", fgColor="FFC7CE")
    B_FILL  = PatternFill("solid", fgColor="D9E1F2")
    H_FONT  = Font(bold=True, color="FFFFFF", size=10)
    BD_FONT = Font(bold=True, size=10)
    NM_FONT = Font(size=10)
    CENTER  = Alignment(horizontal="center")
    THIN    = Side(style="thin")
    BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def apply_header(ws, headers):
        ws.append(headers)
        for ci, _ in enumerate(headers, 1):
            c = ws.cell(1, ci)
            c.fill = H_FILL; c.font = H_FONT
            c.alignment = CENTER; c.border = BORDER

    def style_row(ws, ri, n_cols, fill=None):
        for ci in range(1, n_cols + 1):
            c = ws.cell(ri, ci)
            c.font = NM_FONT; c.border = BORDER; c.alignment = CENTER
        if fill:
            for ci in fill:
                ws.cell(ri, ci).fill = fill[ci]

    def autofit(ws, min_w=10):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=min_w)
            ws.column_dimensions[col[0].column_letter].width = max(w + 2, min_w)

    wb  = Workbook()
    strats = list(STRATEGY_DEFS.keys())

    # ── Sheet 1: Aylık Özet ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Aylık_Özet"
    hdr1 = (["Ay", "SP500_Getiri", "SP500_Deger"] +
            [f"{s}_Getiri" for s in strats] +
            [f"{s}_Deger"  for s in strats])
    apply_header(ws1, hdr1)

    for _, r in monthly.iterrows():
        row = [
            r["ay"].strftime("%Y-%m"),
            pct(r.get("sp500_getiri")),
            usd(r.get("sp500_deger")),
        ]
        for s in strats:
            row.append(pct(r.get(f"{s}_getiri")))
        for s in strats:
            row.append(usd(r.get(f"{s}_deger")))
        ws1.append(row)
        ri = ws1.max_row
        for ci in range(1, len(hdr1) + 1):
            ws1.cell(ri, ci).font = NM_FONT
            ws1.cell(ri, ci).border = BORDER
            ws1.cell(ri, ci).alignment = CENTER
        # Color Bir_Pure getiri column
        bir_a_ci = 3 + strats.index("Bir_Pure") + 1  # 1-indexed
        bir_a_v  = r.get("Bir_Pure_getiri")
        if pd.notna(bir_a_v):
            ws1.cell(ri, bir_a_ci).fill = G_FILL if bir_a_v >= 0 else R_FILL

    for ci in range(1, len(hdr1) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 16

    # ── Sheet 2: Hisse Detay ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Hisse_Detay")
    hdr2 = ["Ay", "Strateji", "Hisse", "Sektör",
             "Rank M1", "Rank M2", "Rank Birleşik",
             "Hot 1m", "Getiri", "Giriş ($)", "Çıkış ($)", "Kâr/Zarar ($)"]
    apply_header(ws2, hdr2)

    for _, r in detail.iterrows():
        kz = r["cikis"] - r["giris"]
        ws2.append([
            r["ay"].strftime("%Y-%m"), r["strateji"], r["hisse"], r.get("sektor",""),
            fmt4(r.get("rank_m1")), fmt4(r.get("rank_m2")), fmt4(r.get("rank_bir")),
            r.get("hot_1m",""),
            pct(r.get("getiri")),
            usd(r.get("giris")), usd(r.get("cikis")), usd(kz),
        ])
        ri   = ws2.max_row
        fill = G_FILL if pd.notna(r.get("getiri")) and r["getiri"] >= 0 else R_FILL
        for ci in range(1, len(hdr2)+1):
            ws2.cell(ri, ci).font   = NM_FONT
            ws2.cell(ri, ci).border = BORDER
            ws2.cell(ri, ci).alignment = CENTER
        if pd.notna(r.get("getiri")):
            ws2.cell(ri, 9).fill = fill

    for ci in range(1, len(hdr2)+1):
        ws2.column_dimensions[get_column_letter(ci)].width = 15

    # ── Sheet 3: Performans Özeti ─────────────────────────────────────────
    ws3 = wb.create_sheet("Performans_Özeti")
    sumdf = _build_summary_df(metrics)
    hdr3  = list(sumdf.columns)
    apply_header(ws3, hdr3)

    for _, r in sumdf.iterrows():
        ws3.append([r[h] for h in hdr3])
        ri = ws3.max_row
        for ci in range(1, len(hdr3)+1):
            ws3.cell(ri, ci).font = NM_FONT
            ws3.cell(ri, ci).border = BORDER
            ws3.cell(ri, ci).alignment = CENTER

    # Model IC bilgisi
    ws3.append([])
    ws3.append(["── Model IC Metrikleri ──"])
    ws3.cell(ws3.max_row, 1).fill = B_FILL
    ws3.cell(ws3.max_row, 1).font = Font(bold=True, size=11)

    def _ic_rows(art, label):
        if art is None:
            return []
        v = art.get("val_ics") or {}
        o = art.get("oot_ics") or {}
        w = art.get("weights", [1/3]*3)
        f = lambda x: f"{x:.4f}" if isinstance(x, float) else "N/A"
        return [
            [f"{label} CV IC — XGBoost",    f(v.get("xgb"))],
            [f"{label} CV IC — LightGBM",   f(v.get("lgb"))],
            [f"{label} CV IC — CatBoost",   f(v.get("cat"))],
            [f"{label} OOT IC — Ensemble",  f(o.get("ensemble"))],
            [f"{label} Holdout IC",         f(art.get("ho_ic"))],
            [f"{label} Top-1 Precision",    f(art.get("top1_precision"))],
            [f"{label} Ağırlık XGB/LGB/CAT",f"{w[0]:.2f}/{w[1]:.2f}/{w[2]:.2f}"],
        ]

    for row in _ic_rows(art1, "Momentum") + _ic_rows(art2, "Reversal"):
        ws3.append(row)
        ri = ws3.max_row
        ws3.cell(ri, 1).font = NM_FONT
        ws3.cell(ri, 1).alignment = CENTER
        if len(row) > 1:
            ws3.cell(ri, 2).font = NM_FONT
            ws3.cell(ri, 2).alignment = CENTER

    ws3.column_dimensions["A"].width = 35
    for ci in range(2, len(hdr3)+1):
        ws3.column_dimensions[get_column_letter(ci)].width = 15

    # ── Sheet 4: Aylık Adaylar ────────────────────────────────────────────
    ws4 = wb.create_sheet("Aylık_Adaylar")
    hdr4 = ["Ay", "Sıra", "Hisse", "Rank M1", "Rank M2",
            "Rank Birleşik", "Gerçek Getiri", "Hot 1m"]
    apply_header(ws4, hdr4)

    prev_ay = None
    for _, r in candidates.iterrows():
        ay_str = r["ay"].strftime("%Y-%m")
        ws4.append([
            ay_str if r["ay"] != prev_ay else "",
            int(r["sira"]), r["hisse"],
            fmt4(r.get("rank_m1")), fmt4(r.get("rank_m2")), fmt4(r.get("rank_bir")),
            pct(r.get("getiri")) if pd.notna(r.get("getiri")) else "",
            r.get("hot_1m",""),
        ])
        ri   = ws4.max_row
        fill = G_FILL if pd.notna(r.get("getiri")) and r["getiri"] >= 0 else R_FILL
        for ci in range(1, len(hdr4)+1):
            ws4.cell(ri, ci).font = NM_FONT
            ws4.cell(ri, ci).border = BORDER
            ws4.cell(ri, ci).alignment = CENTER
        if pd.notna(r.get("getiri")):
            ws4.cell(ri, 7).fill = fill
        prev_ay = r["ay"]

    for ci, w in enumerate([10, 6, 10, 12, 12, 14, 14, 8], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    print(f"Excel kaydedildi: {path}")


# ── Grafik ────────────────────────────────────────────────────────────────
def save_chart(monthly: pd.DataFrame, init_cash: float, path: Path):
    if monthly.empty:
        return

    fig = plt.figure(figsize=(18, 11))
    fig.patch.set_facecolor("#0D1117")
    gs  = GridSpec(2, 2, figure=fig, hspace=0.45, wspace=0.35)

    GOLD   = "#FFD700"; BLUE  = "#4FC3F7"; GREEN = "#69F0AE"
    RED    = "#FF5252"; GRAY  = "#B0BEC5"; BG    = "#161B22"
    PURPLE = "#CE93D8"; AX_C  = "#21262D"

    palette = {
        "Bir_Pure": GOLD,
        "Mom_Pure": BLUE,
        "Reversal": PURPLE,
        "Karma":    GREEN,
        "S&P500":   GRAY,
    }

    dates = pd.to_datetime(monthly["ay"])

    # Panel 1 — Portföy değeri (key strategies + S&P500)
    ax1 = fig.add_subplot(gs[0, :])
    ax1.set_facecolor(AX_C)
    for label, color in palette.items():
        col = "sp500_deger" if label == "S&P500" else f"{label}_deger"
        if col not in monthly.columns:
            continue
        ls = "--" if label == "S&P500" else "-"
        ax1.plot(dates, monthly[col], color=color, lw=1.8, ls=ls, label=label)

    ax1.set_title("Portföy Değeri vs S&P500 ($)", color="white", fontsize=13, pad=10)
    ax1.set_ylabel("$", color=GRAY)
    ax1.tick_params(colors=GRAY)
    ax1.spines[:].set_color("#30363D")
    ax1.legend(facecolor=BG, labelcolor="white", fontsize=9, ncol=5)
    ax1.yaxis.set_major_formatter(mticker.FormatStrFormatter("$%.1f"))

    # Panel 2 — Aylık getiri bar (Bir_Pure)
    ax2 = fig.add_subplot(gs[1, 0])
    ax2.set_facecolor(AX_C)
    if "Bir_Pure_getiri" in monthly.columns:
        rets = monthly["Bir_Pure_getiri"]
        bar_colors = [GREEN if v >= 0 else RED for v in rets]
        ax2.bar(dates, rets * 100, color=bar_colors, width=20, alpha=0.85)
        ax2.axhline(0, color=GRAY, lw=0.8, ls="--")
    ax2.set_title("Aylık Getiri — Bir_Pure (%)", color="white", fontsize=12, pad=8)
    ax2.set_ylabel("%", color=GRAY)
    ax2.tick_params(colors=GRAY)
    ax2.spines[:].set_color("#30363D")

    # Panel 3 — Kümülatif getiri
    ax3 = fig.add_subplot(gs[1, 1])
    ax3.set_facecolor(AX_C)
    for label, color in palette.items():
        ret_col = "sp500_getiri" if label == "S&P500" else f"{label}_getiri"
        if ret_col not in monthly.columns:
            continue
        cum = ((1 + monthly[ret_col].fillna(0)).cumprod() - 1) * 100
        ls  = "--" if label == "S&P500" else "-"
        ax3.plot(dates, cum, color=color, lw=1.8, ls=ls, label=label)
    ax3.axhline(0, color=GRAY, lw=0.5, ls=":")
    ax3.set_title("Kümülatif Getiri (%)", color="white", fontsize=12, pad=8)
    ax3.set_ylabel("%", color=GRAY)
    ax3.tick_params(colors=GRAY)
    ax3.spines[:].set_color("#30363D")
    ax3.legend(facecolor=BG, labelcolor="white", fontsize=9, ncol=3)

    # Suptitle
    bir_a_final = monthly["Bir_Pure_deger"].iloc[-1] if "Bir_Pure_deger" in monthly.columns else init_cash
    total_r = (bir_a_final - init_cash) / init_cash * 100
    fig.suptitle(
        f"US Aylık Model (10 Strateji)  |  ${init_cash:.0f}  →  Bir_Pure: ${bir_a_final:.2f}  ({total_r:+.1f}%)",
        color="white", fontsize=14, fontweight="bold", y=0.98,
    )
    plt.savefig(path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    print(f"Grafik kaydedildi: {path}")


# ── Ana akış ──────────────────────────────────────────────────────────────
def main():
    sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser()
    parser.add_argument("--start",   default=OOT_START_DEFAULT, help="OOT başlangıç tarihi")
    parser.add_argument("--capital", type=float, default=INIT_CASH_DEFAULT, help="Başlangıç sermayesi ($)")
    args = parser.parse_args()

    print("\n=== US Aylık Model Backtest (10 Strateji) ===\n")

    art1 = load_art("ensemble_artifacts.pkl")
    art2 = load_art("ensemble_artifacts_reversal.pkl")
    if art1 is None:
        raise FileNotFoundError("Momentum modeli bulunamadı. Önce retrain.py çalıştırın.")
    print(f"Momentum: OK  |  Reversal: {'OK' if art2 else 'YOK'}")
    print(f"OOT başlangıç: {args.start}  |  Sermaye: ${args.capital:.0f}\n")

    df = load_latest_parquet()
    df = add_ret_ranks(df)

    monthly, detail, candidates = run_backtest(df, art1, art2, args.start, args.capital)

    if monthly.empty:
        print("OOT döneminde yeterli veri yok.")
        return

    metrics = compute_metrics(monthly, args.capital)

    # Konsol özeti
    print(f"\n{'='*62}")
    print(f"  {'Strateji':<12} {'Toplam':>10} {'Yıllık':>10} {'Sharpe':>8} {'Drawdown':>10} {'Alpha':>10}")
    print(f"  {'-'*58}")
    for strat in STRATEGY_DEFS:
        m = metrics.get(strat, {})
        print(f"  {strat:<12} {m.get('Toplam Getiri','N/A'):>10} "
              f"{m.get('Yıllık Getiri','N/A'):>10} "
              f"{m.get('Sharpe','N/A'):>8} "
              f"{m.get('Max Drawdown','N/A'):>10} "
              f"{m.get('Alpha (yıllık)','N/A'):>10}")
    sp = metrics.get("__sp500__", {})
    print(f"  {'S&P500':<12} {sp.get('Toplam Getiri','N/A'):>10} "
          f"{sp.get('Yıllık Getiri','N/A'):>10} {'—':>8} {'—':>10} {'—':>10}")
    print(f"{'='*62}")

    OUTPUT_DIR.mkdir(exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d")
    xlsx = OUTPUT_DIR / f"backtest_{ts}.xlsx"
    png  = OUTPUT_DIR / f"backtest_{ts}.png"

    save_excel(monthly, detail, candidates, metrics, art1, art2, args.capital, xlsx)
    save_chart(monthly, args.capital, png)


if __name__ == "__main__":
    main()
